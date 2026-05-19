import os
from openpyxl import load_workbook
from tkinter import filedialog, messagebox

def abrir_planilha_inicial(excel_path):
    """Abre a planilha e retorna workbook e worksheet ativos."""
    if not excel_path or not os.path.exists(excel_path):
        messagebox.showerror("Erro", "Arquivo Excel não encontrado.")
        return None, None
    try:
        wb = load_workbook(excel_path)
        ws = wb.active
        return wb, ws
    except Exception as e:
        messagebox.showerror("Erro", f"Falha ao abrir planilha: {e}")
        return None, None

def selecionar_arquivo_inicial():
    """Abre o diálogo para selecionar arquivo Excel."""
    p = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
    return p if p else None

def importar_arquivo():
    """Atalho para selecionar arquivo inicial."""
    return selecionar_arquivo_inicial()

def atualizar_dados(ws, linha_atual, lbl_desc=None, entry_id=None, renderizar_imagem=None, img_label=None):
    """Atualiza dados da interface a partir da worksheet e linha."""
    if ws is None:
        return ""
    val = ws[f"A{linha_atual}"].value
    dado = str(val) if val else ""
    desc = str(ws[f"B{linha_atual}"].value or "")
    setor = str(ws[f"C{linha_atual}"].value or "")
    if lbl_desc:
        lbl_desc.configure(text=f"{desc}\n{setor}")
    if entry_id:
        entry_id.delete(0, 'end')
        entry_id.insert(0, dado)
    if dado and renderizar_imagem and img_label:
        img = renderizar_imagem(dado)
        from PIL import ImageTk
        img_tk = ImageTk.PhotoImage(img.resize((500, 160)))
        img_label.configure(image=img_tk)
        img_label.image = img_tk
    return dado
