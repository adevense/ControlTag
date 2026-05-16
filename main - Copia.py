import tkinter as tk
from tkinter import messagebox, filedialog
import customtkinter as ctk
from PIL import Image, ImageDraw, ImageFont, ImageTk
import os
from openpyxl import load_workbook
import pdf417gen

# Bibliotecas para PDF
from reportlab.pdfgen import canvas
from reportlab.lib.units import mm

# Configurações de Aparência
ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("blue")

BASE_PATH = os.path.dirname(os.path.abspath(__file__))
IMG_TEMP_PATH = os.path.join(BASE_PATH, 'codigo_barra.png')
CONFIG_FILE = os.path.join(BASE_PATH, 'config.txt')

class EtiquetaApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Gerador de Etiquetas TIC - Precisão 0.3mm")
        self.geometry("600x600")
        
        self.excel_path = self.carregar_config()
        self.linha_atual = 2
        self.wb = None
        self.ws = None

        self.container = ctk.CTkFrame(self, fg_color="transparent")
        self.container.pack(fill="both", expand=True)
        self.mostrar_menu_principal()

    def carregar_config(self):
        if os.path.exists(CONFIG_FILE):
            with open(CONFIG_FILE, 'r') as f:
                path = f.read().strip()
                return path if os.path.exists(path) else None
        return None

    def salvar_config(self, path):
        with open(CONFIG_FILE, 'w') as f:
            f.write(path)

    def limpar_tela(self):
        for widget in self.container.winfo_children():
            widget.destroy()

    def mostrar_menu_principal(self):
        self.limpar_tela()
        lbl = ctk.CTkLabel(self.container, text="SISTEMA DE ETIQUETAS", font=ctk.CTkFont(size=22, weight="bold"))
        lbl.pack(pady=(50, 30))
        btn_frame = ctk.CTkFrame(self.container, fg_color="transparent")
        btn_frame.pack(expand=True)
        
        ctk.CTkButton(btn_frame, text="📂 Selecionar Planilha Excel", height=55, width=320, 
                      command=self.importar_arquivo).pack(pady=10)

        estado_btn = "normal" if self.excel_path else "disabled"
        ctk.CTkButton(btn_frame, text="🏷️ Ir para Gerador", height=55, width=320, 
                      fg_color="#2c3e50" if self.excel_path else "gray",
                      state=estado_btn, command=self.abrir_gerador).pack(pady=10)

    def importar_arquivo(self):
        path = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx *.xls")])
        if path:
            self.excel_path = path
            self.salvar_config(path)
            self.mostrar_menu_principal()

    def abrir_gerador(self):
        try:
            self.wb = load_workbook(self.excel_path)
            self.ws = self.wb.active
            self.setup_ui_gerador()
            self.atualizar_dados()
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao abrir arquivo: {e}")

    def setup_ui_gerador(self):
        self.limpar_tela()
        header = ctk.CTkFrame(self.container, fg_color="transparent")
        header.pack(fill="x", padx=15, pady=10)
        ctk.CTkButton(header, text="← Menu", width=80, command=self.mostrar_menu_principal).pack(side="left")

        main_card = ctk.CTkFrame(self.container, corner_radius=15)
        main_card.pack(pady=10, padx=20, fill="both", expand=True)

        ctk.CTkLabel(main_card, text="PREVIEW DA ETIQUETA (45x15mm)", font=ctk.CTkFont(weight="bold")).pack(pady=10)

        self.preview_frame = ctk.CTkFrame(main_card, fg_color="white", width=360, height=120)
        self.preview_frame.pack(pady=10)
        self.preview_frame.pack_propagate(False)
        self.img_label = tk.Label(self.preview_frame, bg="white")
        self.img_label.pack(expand=True)

        self.lbl_info = ctk.CTkLabel(main_card, text="-", font=ctk.CTkFont(size=18, weight="bold"))
        self.lbl_info.pack(pady=5)

        nav_frame = ctk.CTkFrame(main_card, fg_color="transparent")
        nav_frame.pack(pady=10)
        self.btn_ant = ctk.CTkButton(nav_frame, text="◀", width=60, command=self.anterior)
        self.btn_ant.grid(row=0, column=0, padx=10)
        self.btn_prox = ctk.CTkButton(nav_frame, text="▶", width=60, command=self.proximo)
        self.btn_prox.grid(row=0, column=1, padx=10)

        ctk.CTkButton(main_card, text="📄 EXPORTAR PDF DUPLO (95.3x15mm)", height=45, fg_color="#e67e22", 
                      command=self.exportar_pdf).pack(pady=5, padx=50, fill="x")
        
        ctk.CTkButton(main_card, text="🖼️ SALVAR IMAGEM PNG", height=45, fg_color="#34495e", 
                      command=self.salvar_etiqueta).pack(pady=5, padx=50, fill="x")

    def exportar_pdf(self):
        if not self.dado: return
        save_path = filedialog.asksaveasfilename(defaultextension=".pdf", 
                                                 initialfile=f"Etiqueta_{self.dado}.pdf",
                                                 filetypes=[("PDF files", "*.pdf")])
        if save_path:
            try:
                # MEDIDAS EXATAS SOLICITADAS
                borda_mm = 2.5 * mm
                entre_etiquetas_mm = 0.3 * mm  # Espaço de 0.3mm
                etiqueta_w_mm = 45 * mm
                altura_mm = 15 * mm
                
                # Largura Total: (2.5 * 2) + (45 * 2) + 0.3 = 95.3mm
                largura_total_mm = (borda_mm * 2) + (etiqueta_w_mm * 2) + entre_etiquetas_mm
                
                c = canvas.Canvas(save_path, pagesize=(largura_total_mm, altura_mm))
                
                # Posição Etiqueta 1 (Esquerda)
                x1 = borda_mm
                c.drawImage(IMG_TEMP_PATH, x1, 0, width=etiqueta_w_mm, height=altura_mm)
                
                # Posição Etiqueta 2 (Direita)
                x2 = borda_mm + etiqueta_w_mm + entre_etiquetas_mm
                c.drawImage(IMG_TEMP_PATH, x2, 0, width=etiqueta_w_mm, height=altura_mm)
                
                c.showPage()
                c.save()
                messagebox.showinfo("Sucesso", "PDF Gerado com espaçamento de 0.3mm!")
                os.startfile(save_path)
            except Exception as e:
                messagebox.showerror("Erro", f"Erro no PDF: {e}")

    def salvar_etiqueta(self):
        if not self.dado: return
        save_path = filedialog.asksaveasfilename(defaultextension=".png",
                                                 initialfile=f"Etiqueta_{self.dado}.png",
                                                 filetypes=[("PNG files", "*.png")])
        if save_path:
            Image.open(IMG_TEMP_PATH).save(save_path)
            messagebox.showinfo("Sucesso", "Imagem salva!")

    def atualizar_dados(self):
        valor = self.ws[f"A{self.linha_atual}"].value
        self.dado = str(valor) if valor else ""
        self.lbl_info.configure(text=f"Patrimônio: {self.dado}")
        self.gerar_etiqueta_imagem()
        self.btn_ant.configure(state="normal" if self.linha_atual > 2 else "disabled")

    def proximo(self):
        self.linha_atual += 1
        if self.ws[f"A{self.linha_atual}"].value: self.atualizar_dados()
        else: self.linha_atual -= 1

    def anterior(self):
        if self.linha_atual > 2:
            self.linha_atual -= 1
            self.atualizar_dados()

    def gerar_etiqueta_imagem(self):
        if not self.dado: return
        # Proporção 3:1 para 45x15mm
        largura, altura = 900, 300 
        barcode_codes = pdf417gen.encode(self.dado, columns=6)
        barcode_img = pdf417gen.render_image(barcode_codes, scale=5, ratio=3, padding=1)
        barcode_img = barcode_img.resize((700, 150), Image.Resampling.LANCZOS)
        
        etiqueta = Image.new("RGB", (largura, altura), "white")
        draw = ImageDraw.Draw(etiqueta)
        try:
            f_g = ImageFont.truetype("arial.ttf", 55)
            f_p = ImageFont.truetype("arial.ttf", 50)
        except:
            f_g = f_p = ImageFont.load_default()
        
        draw.text(((largura-draw.textbbox((0,0),"GESTÃO DE ATIVOS TIC",font=f_g)[2])//2, 20), "GESTÃO DE ATIVOS TIC", fill="black", font=f_g)
        etiqueta.paste(barcode_img, ((largura-700)//2, 85))
        draw.text(((largura-draw.textbbox((0,0),self.dado,font=f_p)[2])//2, 240), self.dado, fill="black", font=f_p)
        
        etiqueta.save(IMG_TEMP_PATH)
        img_tk = ImageTk.PhotoImage(etiqueta.resize((360, 120)))
        self.img_label.configure(image=img_tk); self.img_label.image = img_tk

if __name__ == "__main__":
    app = EtiquetaApp()
    app.mainloop()