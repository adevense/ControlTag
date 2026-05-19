import tkinter as tk
from tkinter import messagebox, filedialog, ttk
import customtkinter as ctk
from PIL import Image, ImageDraw, ImageFont, ImageTk
import os, sys, json, shutil
from core import backup
from config.translations import TEXTS
from config.themes import THEMES
from datetime import datetime
from openpyxl import load_workbook
import pdf417gen
from reportlab.pdfgen import canvas
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader

# --- IMPRESSÃO DIRETA (Windows) ---
try:
    import win32print
    import win32api
    HAS_WIN32 = True
except ImportError:
    HAS_WIN32 = False

# --- FUNÇÃO PARA ENCONTRAR ARQUIVOS NO EXECUTÁVEL ---
def resource_path(relative_path):
    """ Obtém o caminho absoluto para o recurso, funciona para dev e para PyInstaller """
    try:
        # PyInstaller cria uma pasta temporária e armazena o caminho em _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

# --- CONFIGURAÇÃO DE CAMINHOS ---
APP_NAME = "ControlTag_BDGC"
DEFAULT_DATA_DIR = os.path.join(os.environ["APPDATA"], APP_NAME)

if not os.path.exists(DEFAULT_DATA_DIR):
    os.makedirs(DEFAULT_DATA_DIR)

CONFIG_FILE = os.path.join(DEFAULT_DATA_DIR, 'settings.json')
DEFAULT_BACKUP_DIR = os.path.join(DEFAULT_DATA_DIR, 'backups')

if not os.path.exists(DEFAULT_BACKUP_DIR):
    os.makedirs(DEFAULT_BACKUP_DIR)



class EtiquetaApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Control Tag BDGC v1.0")
        self.geometry("1366x900")
        
        # --- DEFINIR ÍCONE DA JANELA ---
        # Tenta carregar o icon.ico se existir
        icon_path = resource_path(os.path.join("resources", "icon.ico"))
        if os.path.exists(icon_path):
            try:
                self.iconbitmap(icon_path)
            except:
                pass # Ignora se der erro no ícone (ex: Linux/Mac ou formato inválido)
        
        # --- CARREGAR CONFIG ---
        self.config = self.carregar_config()
        self.excel_path = self.config.get("last_file")
        self.linha_atual = self.config.get("last_line", 2)
        
        # Pasta de Backup
        self.backup_dir = self.config.get("backup_path", DEFAULT_BACKUP_DIR)
        if not os.path.exists(self.backup_dir):
            try: os.makedirs(self.backup_dir)
            except: self.backup_dir = DEFAULT_BACKUP_DIR
        
        # Correção de Tema
        saved_theme = self.config.get("theme_name", "Padrão (Azul Tech)")
        self.tema_atual = saved_theme if saved_theme in THEMES else "Padrão (Azul Tech)"
        
        self.idioma = self.config.get("language", "pt")
        if self.idioma not in TEXTS: self.idioma = "pt"

        self.escala_ui = self.config.get("ui_scale", 1.0)
        
        self.print_cfg = self.config.get("print_config", {
            "width": 45.0, "height": 15.0, "margin_x": 2.5, "gap": 0.3, "offset_x": 0.0, "offset_y": 0.0
        })
        self.target_printer = self.config.get("target_printer", "")
        self.direct_print_mode = self.config.get("direct_print", False)

        self.fila_impressao = [] 
        ctk.set_widget_scaling(self.escala_ui)

        # --- LAYOUT ---
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(0, weight=1)

        self.setup_sidebar()
        self.main_container = ctk.CTkFrame(self, corner_radius=0)
        self.main_container.grid(row=0, column=1, sticky="nsew")
        self.main_container.grid_rowconfigure(0, weight=1)
        self.main_container.grid_columnconfigure(0, weight=1)
        
        self.frames = {}
        self.setup_pages()

        if self.excel_path and os.path.exists(self.excel_path):
            self.abrir_planilha_inicial()
        else:
            self.status_msg("Aguardando arquivo...")

        self.show_frame("Gerador")
        self.aplicar_tema_visual(self.tema_atual)
        
        self.bind('<Left>', lambda e: self.anterior())
        self.bind('<Right>', lambda e: self.proximo())
        self.bind('<Return>', lambda e: self.buscar_patrimonio() if self.focus_get() == self.entry_busca else None)
        
        self.atualizar_textos_idioma()

    def status_msg(self, msg):
        if hasattr(self, 'status_bar'):
            self.status_bar.configure(text=msg)
            self.update_idletasks()

    def setup_sidebar(self):
        self.sidebar = ctk.CTkFrame(self, width=280, corner_radius=0)
        self.sidebar.grid(row=0, column=0, sticky="nsew")
        self.sidebar.grid_rowconfigure(6, weight=1)

        # --- LOGO (IMAGEM OU TEXTO) ---
        img_path = resource_path(os.path.join("resources", "logo.png"))
        if os.path.exists(img_path):
            try:
                # Carrega imagem se existir
                pil_image = Image.open(img_path)
                # Cria CTkImage (tamanho 150x150, ajustável)
                self.logo_image = ctk.CTkImage(light_image=pil_image, dark_image=pil_image, size=(160, 160))
                self.lbl_logo = ctk.CTkLabel(self.sidebar, text="", image=self.logo_image)
                self.lbl_logo.pack(pady=(40, 5))
            except Exception as e:
                # Fallback para texto se der erro na imagem
                self.lbl_logo = ctk.CTkLabel(self.sidebar, text="CONTROL\nTAG BDGC", font=("Impact", 32))
                self.lbl_logo.pack(pady=(40, 5))
        else:
             # Fallback para texto se arquivo não existir
            self.lbl_logo = ctk.CTkLabel(self.sidebar, text="CONTROL\nTAG BDGC", font=("Impact", 32))
            self.lbl_logo.pack(pady=(40, 5))

        self.lbl_version = ctk.CTkLabel(self.sidebar, text="v1.0", font=("Arial", 12))
        self.lbl_version.pack(pady=(0, 30))

        self.btn_nav_gerador = self.criar_botao_menu("nav_0", "Gerador")
        self.btn_nav_tabela  = self.criar_botao_menu("nav_1", "Tabela")
        self.btn_nav_fila    = self.criar_botao_menu("nav_2", "Fila")
        self.btn_nav_config  = self.criar_botao_menu("nav_3", "Config")

        self.status_bar = ctk.CTkLabel(self.sidebar, text="System Ready", font=("Arial", 11, "italic"), text_color="gray")
        self.status_bar.pack(side="bottom", pady=(5, 10))

        ctk.CTkFrame(self.sidebar, height=1, fg_color="gray").pack(side="bottom", fill="x", padx=20, pady=5)
        self.lbl_credits = ctk.CTkLabel(self.sidebar, text="Dev. Inácio Ribeiro Azevedo", font=("Arial", 10), text_color="gray")
        self.lbl_credits.pack(side="bottom", pady=(0, 15))

    def criar_botao_menu(self, texto_dummy, nome_frame):
        btn = ctk.CTkButton(self.sidebar, text="...", height=50, corner_radius=8,
                            fg_color="transparent", anchor="w", font=("Arial", 15, "bold"),
                            command=lambda: self.show_frame(nome_frame))
        btn.pack(fill="x", padx=15, pady=5)
        return btn

    def setup_pages(self):
        for PageName in ["Gerador", "Tabela", "Fila", "Config"]:
            frame = ctk.CTkFrame(self.main_container, fg_color="transparent")
            self.frames[PageName] = frame
            frame.grid(row=0, column=0, sticky="nsew")
        
        self.setup_page_gerador(self.frames["Gerador"])
        self.setup_page_tabela(self.frames["Tabela"])
        self.setup_page_fila(self.frames["Fila"])
        self.setup_page_config(self.frames["Config"])

    def setup_page_gerador(self, parent):
        parent.grid_columnconfigure(0, weight=2)
        parent.grid_columnconfigure(1, weight=1)
        
        left = ctk.CTkFrame(parent, fg_color="transparent")
        left.grid(row=0, column=0, sticky="nsew", padx=20, pady=20)
        
        self.lbl_preview_title = ctk.CTkLabel(left, text="Preview", font=("Arial", 12, "bold"))
        self.lbl_preview_title.pack(anchor="w")
        
        self.preview_frame = ctk.CTkFrame(left, fg_color="white", height=250)
        self.preview_frame.pack(fill="x", pady=5)
        self.preview_frame.pack_propagate(False)
        self.img_label = tk.Label(self.preview_frame, bg="white")
        self.img_label.pack(expand=True, fill="both")

        card = ctk.CTkFrame(left)
        card.pack(fill="both", expand=True, pady=20)
        
        self.lbl_desc_title = ctk.CTkLabel(card, text="Description", font=("Arial", 16, "bold"))
        self.lbl_desc_title.pack(pady=10)
        self.lbl_desc = ctk.CTkLabel(card, text="--", font=("Arial", 20), wraplength=500)
        self.lbl_desc.pack(pady=5)

        nav = ctk.CTkFrame(card, fg_color="transparent")
        nav.pack(side="bottom", pady=30)
        self.btn_ant = ctk.CTkButton(nav, text="<", width=120, height=45, command=self.anterior)
        self.btn_ant.pack(side="left", padx=15)
        self.entry_id = ctk.CTkEntry(nav, width=200, height=45, font=("Arial", 22, "bold"), justify="center")
        self.entry_id.pack(side="left", padx=10)
        self.btn_prox = ctk.CTkButton(nav, text=">", width=120, height=45, command=self.proximo)
        self.btn_prox.pack(side="left", padx=15)
        self.btn_save = ctk.CTkButton(nav, text="Save", width=80, fg_color="#555", command=self.salvar_edicao)
        self.btn_save.pack(side="bottom", pady=5)

        right = ctk.CTkFrame(parent)
        right.grid(row=0, column=1, sticky="nsew")
        
        self.lbl_controls_title = ctk.CTkLabel(right, text="Controls", font=("Arial", 18, "bold"))
        self.lbl_controls_title.pack(pady=20)
        
        self.entry_busca = ctk.CTkEntry(right, placeholder_text="...")
        self.entry_busca.pack(fill="x", padx=20, pady=(0, 5))
        self.btn_buscar = ctk.CTkButton(right, text="Search", command=self.buscar_patrimonio)
        self.btn_buscar.pack(fill="x", padx=20, pady=(0, 30))
        
        self.btn_add = ctk.CTkButton(right, text="Add Queue", height=50, command=self.add_fila)
        self.btn_add.pack(fill="x", padx=20, pady=10)
        self.btn_pdf_unico = ctk.CTkButton(right, text="PDF Single", fg_color="#444", command=self.exportar_pdf_unico)
        self.btn_pdf_unico.pack(fill="x", padx=20, pady=10)
        
        ctk.CTkFrame(right, height=2, fg_color="gray").pack(fill="x", padx=20, pady=20)
        self.lbl_batch_title = ctk.CTkLabel(right, text="Batch", font=("Arial", 14))
        self.lbl_batch_title.pack()
        self.batch_start = ctk.CTkEntry(right, justify="center")
        self.batch_start.pack(pady=5)
        self.batch_end = ctk.CTkEntry(right, justify="center")
        self.batch_end.pack(pady=5)
        self.btn_lote = ctk.CTkButton(right, text="Batch", fg_color="#c0392b", command=self.gerar_lote_intervalo)
        self.btn_lote.pack(fill="x", padx=20, pady=10)

    def setup_page_tabela(self, parent):
        frame_tree = ctk.CTkFrame(parent)
        frame_tree.pack(fill="both", expand=True, padx=20, pady=20)
        cols = ("Linha", "Patrimônio", "Descrição", "Setor")
        self.tree = ttk.Treeview(frame_tree, columns=cols, show="headings")
        self.tree.column("Linha", width=50); self.tree.column("Patrimônio", width=150)
        self.tree.column("Descrição", width=400); self.tree.column("Setor", width=200)
        sb = ttk.Scrollbar(frame_tree, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscroll=sb.set)
        self.tree.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")
        self.tree.bind("<Double-1>", self.selecionar_na_tabela)
        self.btn_refresh = ctk.CTkButton(parent, text="Refresh", command=self.carregar_dados_tabela)
        self.btn_refresh.pack(pady=10)

    def setup_page_fila(self, parent):
        self.lbl_queue_count = ctk.CTkLabel(parent, text="0", font=("Arial", 20))
        self.lbl_queue_count.pack(pady=20)
        self.lista_fila = ctk.CTkTextbox(parent, width=700, height=400, font=("Consolas", 14))
        self.lista_fila.pack(pady=10)
        btn_frame = ctk.CTkFrame(parent, fg_color="transparent")
        btn_frame.pack(pady=20)
        self.btn_clear = ctk.CTkButton(btn_frame, text="Clear", fg_color="red", command=self.limpar_fila)
        self.btn_clear.pack(side="left", padx=10)
        self.btn_print_queue = ctk.CTkButton(btn_frame, text="Print Queue", width=200, command=self.gerar_pdf_fila)
        self.btn_print_queue.pack(side="left", padx=10)

    def setup_page_config(self, parent):
        self.tabview = ctk.CTkTabview(parent, width=800, height=600)
        self.tabview.pack(pady=20, padx=20, fill="both", expand=True)
        
        self.tab_geral = self.tabview.add("Geral")
        self.tab_visual = self.tabview.add("Aparência")
        self.tab_print = self.tabview.add("Impressão")

        # --- TAB 1: GERAL ---
        self.lbl_conf_lang = ctk.CTkLabel(self.tab_geral, text="Language", font=("Arial", 14, "bold"))
        self.lbl_conf_lang.pack(pady=(15, 5))
        self.combo_lang = ctk.CTkOptionMenu(self.tab_geral, values=["pt", "en", "es"], command=self.mudar_idioma)
        self.combo_lang.set(self.idioma)
        self.combo_lang.pack(pady=5)

        self.lbl_conf_file = ctk.CTkLabel(self.tab_geral, text="File", font=("Arial", 14, "bold"))
        self.lbl_conf_file.pack(pady=(15, 5))
        self.btn_file = ctk.CTkButton(self.tab_geral, text="Select File", command=self.importar_arquivo)
        self.btn_file.pack(pady=5)

        # --- SEÇÃO DE BACKUP ---
        self.lbl_conf_extra = ctk.CTkLabel(self.tab_geral, text="Backup", font=("Arial", 14, "bold"))
        self.lbl_conf_extra.pack(pady=(20, 5))
        
        self.lbl_path_backup_title = ctk.CTkLabel(self.tab_geral, text="Pasta Atual:", font=("Arial", 10))
        self.lbl_path_backup_title.pack()
        self.lbl_path_backup_val = ctk.CTkLabel(self.tab_geral, text=self.backup_dir, font=("Arial", 10, "italic"), text_color="gray")
        self.lbl_path_backup_val.pack(pady=2)

        frame_bkp = ctk.CTkFrame(self.tab_geral, fg_color="transparent")
        frame_bkp.pack(pady=10)
        
        self.btn_change_bkp = ctk.CTkButton(frame_bkp, text="Mudar Pasta", width=140, fg_color="#555", command=self.selecionar_pasta_backup)
        self.btn_change_bkp.grid(row=0, column=0, padx=5, pady=5)
        
        self.btn_backup = ctk.CTkButton(frame_bkp, text="Fazer Backup", width=140, fg_color="green", command=self.fazer_backup_manual)
        self.btn_backup.grid(row=0, column=1, padx=5, pady=5)
        
        self.btn_open_folder = ctk.CTkButton(frame_bkp, text="Abrir Pasta", width=140, fg_color="#333", command=self.abrir_pasta_dados)
        self.btn_open_folder.grid(row=1, column=0, padx=5, pady=5)
        
        self.btn_clear_bkp = ctk.CTkButton(frame_bkp, text="Limpar Backups", width=140, fg_color="#c0392b", hover_color="#922b21", command=self.limpar_backups)
        self.btn_clear_bkp.grid(row=1, column=1, padx=5, pady=5)

        # --- TAB 2: APARÊNCIA ---
        self.lbl_conf_theme = ctk.CTkLabel(self.tab_visual, text="Theme", font=("Arial", 14, "bold"))
        self.lbl_conf_theme.pack(pady=(20, 5))
        self.combo_tema = ctk.CTkOptionMenu(self.tab_visual, values=list(THEMES.keys()), command=self.mudar_tema_callback)
        self.combo_tema.set(self.tema_atual)
        self.combo_tema.pack(pady=5)

        self.lbl_conf_scale = ctk.CTkLabel(self.tab_visual, text="Scale", font=("Arial", 14, "bold"))
        self.lbl_conf_scale.pack(pady=(20, 5))
        self.slider = ctk.CTkSlider(self.tab_visual, from_=0.8, to=1.4, command=self.mudar_escala)
        self.slider.set(self.escala_ui)
        self.slider.pack(pady=5)

        # --- TAB 3: IMPRESSÃO ---
        frame_hardware = ctk.CTkFrame(self.tab_print, fg_color="transparent")
        frame_hardware.pack(fill="x", padx=20, pady=(20, 10))
        
        self.lbl_sel_print = ctk.CTkLabel(frame_hardware, text="Selecionar Impressora", font=("Arial", 12, "bold"))
        self.lbl_sel_print.pack(anchor="w")
        
        lista_impressoras = self.listar_impressoras_sistema()
        self.combo_printer = ctk.CTkOptionMenu(frame_hardware, values=lista_impressoras, command=self.selecionar_impressora)
        if self.target_printer in lista_impressoras:
            self.combo_printer.set(self.target_printer)
        else:
            self.combo_printer.set("Selecione...")
        self.combo_printer.pack(fill="x", pady=5)

        self.switch_direct = ctk.CTkSwitch(frame_hardware, text="Modo Direto", command=self.toggle_impressao_direta)
        if self.direct_print_mode: self.switch_direct.select()
        else: self.switch_direct.deselect()
        self.switch_direct.pack(pady=10, anchor="w")

        ctk.CTkLabel(self.tab_print, text="--------------------------------").pack()
        ctk.CTkLabel(self.tab_print, text="Ajuste Dimensional (mm)", text_color="orange").pack(pady=10)
        
        self.entries_print = {}
        keys = ["width", "height", "margin_x", "gap", "offset_x", "offset_y"]
        grid_frame = ctk.CTkFrame(self.tab_print, fg_color="transparent")
        grid_frame.pack(pady=10)

        for i, k in enumerate(keys):
            lbl = ctk.CTkLabel(grid_frame, text=k)
            lbl.grid(row=i, column=0, padx=10, pady=5, sticky="e")
            entry = ctk.CTkEntry(grid_frame, width=80, justify="center")
            entry.insert(0, str(self.print_cfg[k]))
            entry.grid(row=i, column=1, padx=10, pady=5)
            self.entries_print[k] = (lbl, entry)

        btn_save_print = ctk.CTkButton(self.tab_print, text="💾 Salvar Config", command=self.salvar_cfg_impressao)
        btn_save_print.pack(pady=20)
        
        self.btn_reset_print = ctk.CTkButton(self.tab_print, text="Reset Defaults", fg_color="#444", width=100, command=self.reset_print_cfg)
        self.btn_reset_print.pack()

    def listar_impressoras_sistema(self):
        if not HAS_WIN32: return ["PyWin32 não instalado"]
        try:
            printers = win32print.EnumPrinters(win32print.PRINTER_ENUM_LOCAL | win32print.PRINTER_ENUM_CONNECTIONS)
            return [p[2] for p in printers]
        except Exception as e:
            return [f"Erro: {str(e)}"]

    def selecionar_impressora(self, p_name):
        self.target_printer = p_name
        self.salvar_config()

    def toggle_impressao_direta(self):
        self.direct_print_mode = bool(self.switch_direct.get())
        self.salvar_config()

    def enviar_arquivo_para_impressora(self, filepath):
        if not HAS_WIN32 or not self.direct_print_mode:
            os.startfile(filepath)
            return

        if not self.target_printer or "Erro" in self.target_printer or "Selecione" in self.target_printer:
            messagebox.showwarning("Impressora", "Selecione uma impressora válida na aba Config!")
            os.startfile(filepath)
            return

        try:
            win32api.ShellExecute(0, "printto", filepath, f'"{self.target_printer}"', ".", 0)
            self.status_msg(f"Enviado para: {self.target_printer}")
        except Exception as e:
            messagebox.showerror("Erro de Impressão", f"Falha ao enviar: {e}\nAbrindo PDF...")
            os.startfile(filepath)
    
    # --- FUNÇÕES DE BACKUP ---
    def abrir_pasta_dados(self):
        backup.abrir_pasta_dados(self.backup_dir)

    def selecionar_pasta_backup(self):
        path = backup.selecionar_pasta_backup(self.backup_dir, filedialog, lambda p: self.lbl_path_backup_val.configure(text=p))
        if path:
            self.backup_dir = path
            self.salvar_config()
            messagebox.showinfo("Sucesso", f"Backups agora serão salvos em:\n{path}")

    def limpar_backups(self):
        backup.limpar_backups(self.backup_dir, self.idioma, TEXTS, self.status_msg)

    def fazer_backup_manual(self):
        backup.fazer_backup_manual(self.excel_path, self.backup_dir)
    
    # --- FIM FUNÇÕES BACKUP ---

    def mudar_idioma(self, novo_lang):
        self.idioma = novo_lang
        self.atualizar_textos_idioma()
        self.salvar_config()
        self.status_msg(f"Language: {novo_lang.upper()}")

    def atualizar_textos_idioma(self):
        t = TEXTS[self.idioma]
        btns = [self.btn_nav_gerador, self.btn_nav_tabela, self.btn_nav_fila, self.btn_nav_config]
        for btn, txt in zip(btns, t["nav"]): btn.configure(text=txt)
        self.lbl_credits.configure(text=t["credits"])
        
        self.lbl_preview_title.configure(text=t["ui"][0])
        self.lbl_desc_title.configure(text=t["ui"][1])
        self.lbl_controls_title.configure(text=t["ui"][2])
        self.btn_ant.configure(text=t["ui"][3])
        self.btn_prox.configure(text=t["ui"][4])
        self.btn_save.configure(text=t["ui"][5])
        
        self.entry_busca.configure(placeholder_text=t["actions"][0])
        self.btn_buscar.configure(text=t["actions"][1])
        self.btn_add.configure(text=t["actions"][2])
        self.btn_pdf_unico.configure(text=t["actions"][3])
        self.batch_start.configure(placeholder_text=t["actions"][4])
        self.batch_end.configure(placeholder_text=t["actions"][5])
        self.btn_lote.configure(text=t["actions"][6])
        
        self.tabview._segmented_button.configure(values=t["tabs"])
        self.lbl_conf_lang.configure(text=t["lbls"][0])
        self.lbl_conf_theme.configure(text=t["lbls"][1])
        self.lbl_conf_scale.configure(text=t["lbls"][2])
        self.lbl_conf_file.configure(text=t["lbls"][3])
        self.lbl_conf_extra.configure(text=t["lbls"][4])
        self.lbl_sel_print.configure(text=t["lbls"][5])
        self.switch_direct.configure(text=t["lbls"][6])
        self.lbl_path_backup_title.configure(text=t["lbls"][7])

        self.btn_reset_print.configure(text=t["msg"][6])
        
        self.btn_change_bkp.configure(text=t["btns"][0])
        self.btn_clear_bkp.configure(text=t["btns"][1])
        self.btn_backup.configure(text=t["btns"][2])
        self.btn_open_folder.configure(text=t["btns"][3])
        
        keys_order = ["width", "height", "margin_x", "gap", "offset_x", "offset_y"]
        for k, txt in zip(keys_order, t["print_cfg"]):
            self.entries_print[k][0].configure(text=txt)

    def salvar_cfg_impressao(self):
        try:
            for k, (_, entry) in self.entries_print.items():
                val = float(entry.get().replace(",", "."))
                self.print_cfg[k] = val
            self.salvar_config()
            self.status_msg(TEXTS[self.idioma]["msg"][1])
        except ValueError:
            messagebox.showerror("Erro", "Números inválidos")

    def reset_print_cfg(self):
        defaults = {"width": 45.0, "height": 15.0, "margin_x": 2.5, "gap": 0.3, "offset_x": 0.0, "offset_y": 0.0}
        self.print_cfg = defaults
        for k, (_, entry) in self.entries_print.items():
            entry.delete(0, 'end'); entry.insert(0, str(defaults[k]))
        self.salvar_config()

    def renderizar_imagem(self, valor):
        largura, altura = 900, 300 
        barcode = pdf417gen.render_image(pdf417gen.encode(valor, columns=6), scale=5, ratio=3, padding=1)
        barcode = barcode.resize((700, 140), Image.Resampling.LANCZOS)
        etiqueta = Image.new("RGB", (largura, altura), "white")
        draw = ImageDraw.Draw(etiqueta)
        try: font = ImageFont.truetype("arial.ttf", 55); font_p = ImageFont.truetype("arial.ttf", 45)
        except: font = font_p = ImageFont.load_default()
        draw.text(((largura-draw.textbbox((0,0),"GESTÃO DE ATIVOS TIC",font=font)[2])//2, 15), "GESTÃO DE ATIVOS TIC", fill="black", font=font)
        etiqueta.paste(barcode, ((largura-700)//2, 80))
        draw.text(((largura-draw.textbbox((0,0),valor,font=font_p)[2])//2, 235), valor, fill="black", font=font_p)
        return etiqueta

    def gerar_pdf_generico(self, lista_dicts, filepath):
        try:
            W_mm = self.print_cfg["width"]
            H_mm = self.print_cfg["height"]
            Marg_mm = self.print_cfg["margin_x"]
            Gap_mm = self.print_cfg["gap"]
            OffX_mm = self.print_cfg["offset_x"]
            OffY_mm = self.print_cfg["offset_y"]

            w = W_mm * mm; h = H_mm * mm
            margem = Marg_mm * mm; gap = Gap_mm * mm
            off_x = OffX_mm * mm; off_y = OffY_mm * mm

            larg_total = (margem * 2) + (w * 2) + gap
            
            c = canvas.Canvas(filepath, pagesize=(larg_total, h))
            
            for item in lista_dicts:
                img_pil = self.renderizar_imagem(item['id'])
                img_mem = ImageReader(img_pil)
                pos_x1 = margem + off_x
                pos_y = off_y
                pos_x2 = margem + w + gap + off_x
                
                c.drawImage(img_mem, pos_x1, pos_y, width=w, height=h)
                c.drawImage(img_mem, pos_x2, pos_y, width=w, height=h)
                c.showPage()
            
            c.save()
            self.enviar_arquivo_para_impressora(filepath)

        except Exception as e: messagebox.showerror("Erro PDF", str(e))

    def carregar_config(self):
        if os.path.exists(CONFIG_FILE):
            try: return json.load(open(CONFIG_FILE))
            except: pass
        return {"language": "pt"}

    def salvar_config(self):
        dados = {
            "last_file": self.excel_path, "last_line": self.linha_atual,
            "theme_name": self.tema_atual, "language": self.idioma, "ui_scale": self.escala_ui,
            "print_config": self.print_cfg,
            "target_printer": self.target_printer,
            "direct_print": self.direct_print_mode,
            "backup_path": self.backup_dir
        }
        with open(CONFIG_FILE, 'w') as f: json.dump(dados, f)

    def abrir_planilha_inicial(self):
        try:
            self.wb = load_workbook(self.excel_path)
            self.ws = self.wb.active
            self.atualizar_dados()
            self.carregar_dados_tabela()
            self.status_msg(f"File Loaded")
        except Exception as e: self.status_msg(f"Erro: {str(e)}")

    def selecionar_arquivo_inicial(self):
        p = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
        if p:
            self.excel_path = p; self.linha_atual = 2
            self.salvar_config(); self.abrir_planilha_inicial()

    def importar_arquivo(self): self.selecionar_arquivo_inicial()

    def atualizar_dados(self):
        if not hasattr(self, 'ws'): return
        val = self.ws[f"A{self.linha_atual}"].value
        self.dado = str(val) if val else ""
        desc = str(self.ws[f"B{self.linha_atual}"].value or "")
        setor = str(self.ws[f"C{self.linha_atual}"].value or "")
        self.lbl_desc.configure(text=f"{desc}\n{setor}")
        self.entry_id.delete(0, 'end'); self.entry_id.insert(0, self.dado)
        if self.dado:
            img = self.renderizar_imagem(self.dado)
            img_tk = ImageTk.PhotoImage(img.resize((500, 160)))
            self.img_label.configure(image=img_tk); self.img_label.image = img_tk

    def salvar_edicao(self):
        if not os.path.exists(self.backup_dir): os.makedirs(self.backup_dir)
        try:
            shutil.copy2(self.excel_path, os.path.join(self.backup_dir, "autobackup.xlsx"))
        except: pass

        novo = self.entry_id.get()
        self.ws[f"A{self.linha_atual}"] = novo
        self.wb.save(self.excel_path)
        self.atualizar_dados(); self.carregar_dados_tabela()
        self.status_msg(TEXTS[self.idioma]["msg"][1])

    def proximo(self): self.linha_atual += 1; self.atualizar_dados()
    def anterior(self): 
        if self.linha_atual > 2: self.linha_atual -= 1; self.atualizar_dados()
    def buscar_patrimonio(self):
        termo = self.entry_busca.get().strip()
        for i in range(2, self.ws.max_row + 1):
            if str(self.ws[f"A{i}"].value).strip() == termo:
                self.linha_atual = i; self.atualizar_dados(); return
        messagebox.showwarning("Busca", "404")

    def add_fila(self):
        item = {"linha": self.linha_atual, "id": self.dado}
        if not any(x['linha'] == self.linha_atual for x in self.fila_impressao):
            self.fila_impressao.append(item)
            self.lista_fila.insert("end", f"▶ {item['id']}\n")
            self.lbl_queue_count.configure(text=str(len(self.fila_impressao)))

    def limpar_fila(self):
        self.fila_impressao = []; self.lista_fila.delete("1.0", "end"); self.lbl_queue_count.configure(text="0")

    def gerar_pdf_fila(self):
        if self.fila_impressao: 
            if self.direct_print_mode:
                p = os.path.join(os.environ["TEMP"], "fila_temp.pdf")
                self.gerar_pdf_generico(self.fila_impressao, p)
            else:
                p = filedialog.asksaveasfilename(defaultextension=".pdf")
                if p: self.gerar_pdf_generico(self.fila_impressao, p)

    def exportar_pdf_unico(self):
        if self.direct_print_mode:
            p = os.path.join(os.environ["TEMP"], f"temp_{self.dado}.pdf")
            self.gerar_pdf_generico([{"linha": self.linha_atual, "id": self.dado}], p)
        else:
            self.gerar_pdf_generico([{"linha": self.linha_atual, "id": self.dado}], f"Tag_{self.dado}.pdf")

    def gerar_lote_intervalo(self):
        try:
            i, f = int(self.batch_start.get()), int(self.batch_end.get())
            lista = [{"linha": r, "id": str(self.ws[f"A{r}"].value)} for r in range(i, f+1) if self.ws[f"A{r}"].value]
            if self.direct_print_mode:
                p = os.path.join(os.environ["TEMP"], "lote_temp.pdf")
                self.gerar_pdf_generico(lista, p)
            else:
                p = filedialog.asksaveasfilename(defaultextension=".pdf")
                if p: self.gerar_pdf_generico(lista, p)
        except: messagebox.showerror("Erro", "Intervalo Inválido")

    def mudar_tema_callback(self, t): self.tema_atual = t; self.aplicar_tema_visual(t); self.salvar_config()
    def mudar_escala(self, v): self.escala_ui = v; ctk.set_widget_scaling(v); self.salvar_config()
    
    def aplicar_tema_visual(self, nome):
        t = THEMES.get(nome, THEMES["Padrão (Azul Tech)"])
        ctk.set_appearance_mode(t["mode"])
        self.sidebar.configure(fg_color=t["sidebar"])
        self.logo_text_color = t["primary"] # Variavel auxiliar
        if hasattr(self, 'lbl_logo') and isinstance(self.lbl_logo, ctk.CTkLabel) and not hasattr(self, 'logo_image'):
            # Se for texto, muda a cor. Se for imagem, não muda.
             self.lbl_logo.configure(text_color=t["primary"])

        for btn in [self.btn_ant, self.btn_prox, self.btn_buscar, self.btn_add, self.btn_lote]:
            btn.configure(fg_color=t["primary"], hover_color=t.get("hover", t["primary"]))
        self.highlight_menu()

    def show_frame(self, p): self.frames[p].tkraise(); self.highlight_menu()
    def highlight_menu(self):
        t = THEMES.get(self.tema_atual)
        if t is None:
            self.tema_atual = "Padrão (Azul Tech)"
            t = THEMES["Padrão (Azul Tech)"]
        for b in [self.btn_nav_gerador, self.btn_nav_tabela, self.btn_nav_fila, self.btn_nav_config]:
            b.configure(text_color=t.get("text_menu", "white"))

    def carregar_dados_tabela(self):
        for i in self.tree.get_children(): self.tree.delete(i)
        if hasattr(self, 'ws'):
            for r in self.ws.iter_rows(min_row=2, max_row=5000, values_only=False):
                if r[0].value: self.tree.insert("", "end", values=(r[0].row, r[0].value, r[1].value, r[2].value))
    
    def selecionar_na_tabela(self, e):
        self.linha_atual = int(self.tree.item(self.tree.selection()[0], "values")[0])
        self.atualizar_dados(); self.show_frame("Gerador")

if __name__ == "__main__":
    app = EtiquetaApp()
    app.mainloop()